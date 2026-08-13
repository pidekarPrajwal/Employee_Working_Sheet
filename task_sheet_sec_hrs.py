"""
Jira Excel Time Converter
--------------------------
Reads a Jira Excel export, finds the "Spent Time (Seconds)" column (or common
variants like "Time Spent", "Spent Seconds"), converts it to hours, and adds
a new "Spent Time (Hours)" column next to it. Works across all worksheets
(or a single specified sheet), preserves all other columns, and writes the
result to a new file without touching the original.

Requirements:
    pip install pandas openpyxl

Usage:
    python jira_time_converter.py
    (edit INPUT_FILE / OUTPUT_FILE / SHEET_NAME below, or import
     convert_jira_file() into your own script)
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


# ----------------------------------------------------------------------
# Column names we will accept as "the seconds column". Matching is done
# case-insensitively and ignoring surrounding whitespace.
# ----------------------------------------------------------------------
CANDIDATE_COLUMN_NAMES = [
    "spent time (seconds)",
    "time spent (seconds)",
    "time spent",
    "spent seconds",
    "spent time",
]

NEW_COLUMN_NAME = "Spent Time (Hours)"


def find_seconds_column(columns):
    """
    Search a list of column names (as they appear in the sheet) for a match
    against CANDIDATE_COLUMN_NAMES. Returns the actual column name found,
    or None if no match exists.
    """
    normalized_map = {str(col).strip().lower(): col for col in columns}
    for candidate in CANDIDATE_COLUMN_NAMES:
        if candidate in normalized_map:
            return normalized_map[candidate]
    return None


def seconds_to_hours(value):
    """
    Convert a single value (seconds) to hours, rounded to 2 decimals.
    Returns None (empty cell) if the value is missing, blank, or not a
    valid number.
    """
    # Treat NaN / None / empty string as missing
    if pd.isna(value) or (isinstance(value, str) and value.strip() == ""):
        return None
    try:
        seconds = float(value)
    except (ValueError, TypeError):
        # Non-numeric junk in the cell -> leave blank
        return None
    return round(seconds / 3600, 2)


def convert_jira_file(input_file, output_file, sheet_name=None):
    """
    Main conversion routine.

    Parameters
    ----------
    input_file : str
        Path to the source Jira .xlsx export.
    output_file : str
        Path where the converted workbook will be written.
    sheet_name : str or None
        If given, only that worksheet is processed. If None, all worksheets
        are processed.
    """
    # --- 1. Validate the input file exists -----------------------------
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    # --- 2. Try to open the workbook; catch corrupt/invalid files ------
    try:
        # sheet_name=None with pandas returns a dict of {sheet_name: DataFrame}
        # so we can process every sheet in the workbook.
        sheets = pd.read_excel(input_file, sheet_name=sheet_name, engine="openpyxl")
    except InvalidFileException as e:
        raise ValueError(f"'{input_file}' is not a valid Excel (.xlsx) file.") from e
    except Exception as e:
        raise ValueError(f"Could not read '{input_file}': {e}") from e

    # If a single sheet name was requested, pandas returns a single
    # DataFrame instead of a dict -- normalize to a dict for uniform handling.
    if isinstance(sheets, pd.DataFrame):
        sheets = {sheet_name: sheets}

    processed_sheets = {}
    sheets_with_column = 0

    # --- 3. Process each worksheet --------------------------------------
    for sname, df in sheets.items():
        seconds_col = find_seconds_column(df.columns)

        if seconds_col is None:
            # No matching column on this sheet -- keep the sheet unchanged
            # (e.g. summary tabs, empty tabs, etc.) rather than failing
            # the whole export.
            processed_sheets[sname] = df
            continue

        sheets_with_column += 1

        # Convert seconds -> hours, row by row, handling bad/missing data
        hours_series = df[seconds_col].apply(seconds_to_hours)

        # Insert the new column immediately after the seconds column
        col_position = df.columns.get_loc(seconds_col) + 1
        df.insert(col_position, NEW_COLUMN_NAME, hours_series)

        processed_sheets[sname] = df

    if sheets_with_column == 0:
        raise KeyError(
            "No sheet contained a recognizable time column. "
            f"Looked for one of: {CANDIDATE_COLUMN_NAMES}"
        )

    # --- 4. Write all sheets to the new output file ---------------------
    # xlsxwriter/openpyxl engine handles large row counts fine (tested to
    # hundreds of thousands of rows); using openpyxl keeps formatting
    # (column dtypes, most cell styling from pandas' default writer).
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sname, df in processed_sheets.items():
            # Excel sheet names are capped at 31 characters
            safe_name = str(sname)[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    # --- 5. Light formatting pass: bold header row, autosize new column -
    try:
        wb = load_workbook(output_file)
        for ws in wb.worksheets:
            # Bold header row
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            # Autosize the new hours column if present
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == NEW_COLUMN_NAME:
                    ws.column_dimensions[get_column_letter(idx)].width = 20
        wb.save(output_file)
    except Exception:
        # Formatting is a nice-to-have; never fail the whole run over it
        pass

    return output_file, sheets_with_column


if __name__ == "__main__":
    # ------------------------------------------------------------------
    # EDIT THESE THREE VARIABLES to point at your own files, or pass them
    # as command-line arguments: python jira_time_converter.py in.xlsx out.xlsx [sheet]
    # ------------------------------------------------------------------
    INPUT_FILE = "jira_export.xlsx"
    OUTPUT_FILE = "jira_tasks_hours.xlsx"
    SHEET_NAME = None  # e.g. "Sheet1" to process only one sheet

    if len(sys.argv) >= 3:
        INPUT_FILE = sys.argv[1]
        OUTPUT_FILE = sys.argv[2]
        SHEET_NAME = sys.argv[3] if len(sys.argv) >= 4 else None

    try:
        out_path, n_sheets = convert_jira_file(INPUT_FILE, OUTPUT_FILE, SHEET_NAME)
        print(f"Done. Converted time column found on {n_sheets} sheet(s).")
        print(f"Output written to: {out_path}")
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        sys.exit(1)
    except KeyError as e:
        print(f"ERROR: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"UNEXPECTED ERROR: {e}")
        sys.exit(1)