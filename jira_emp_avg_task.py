#!/usr/bin/env python3
"""
jira_emp_avg_task.py

Sheet-only report (no leave file, no worklogs file, no Jira API).

Reads a Jira task export and writes one Excel workbook with:
  - one "All Tasks" sheet
  - one sheet per employee (Assignee)

For --month YYYY-MM or --month 07 (default: current calendar month):
  - Include if Created in selected month, OR
    Created in previous month AND Updated in selected month.
  - Uses Σ Time Spent from the sheet as the task time value.
  - Working days = all days except Sundays and ONE Saturday off (default: 2nd Saturday).
    Other Saturdays count as working days.
  - Avg / Working Day = task time ÷ working days
  - Override off Saturday with --saturday-off 1st|2nd|3rd|4th|last

NOTE:
  Jira's "Σ Time Spent" on an issue export is usually LIFETIME total, not
  month-only. This script cannot split June vs August from that column alone.
  If your export was already filtered in Jira for the month, use that file.

Usage:
    python jira_emp_avg_task.py --input "Jira sheet 100.csv"
    python jira_emp_avg_task.py --input "Jira sheet 100.csv" --month 07 --output Emp_Avg.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DURATION_FMT = "[h]:mm"

COLUMN_ALIASES = {
    "Issue Key": ["issue key", "key", "issuekey"],
    "Summary": ["summary"],
    "Assignee": ["assignee"],
    "Created Date": ["created date", "created"],
    "Updated Date": ["updated date", "updated"],
    "Status": ["status"],
    "Priority": ["priority"],
    "Issue Type": ["issue type", "issuetype", "type"],
    "Time Spent": [
        "time spent",
        "\u03a3 time spent",
        "sigma time spent",
        "sum time spent",
        "sum of time spent",
    ],
}


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    lookup = {}
    for canonical, variants in COLUMN_ALIASES.items():
        for v in variants:
            lookup[v.strip().lower()] = canonical
    rename = {}
    for col in df.columns:
        key = re.sub(r"\s+", " ", str(col).strip().lower())
        if key in lookup:
            rename[col] = lookup[key]
            continue
        for variant, canonical in lookup.items():
            if variant in key or key in variant:
                rename[col] = canonical
                break
    return df.rename(columns=rename)


def read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, sheet_name=0, engine="openpyxl")
    if suffix == ".csv":
        return pd.read_csv(path)
    raise SystemExit(f"Unsupported input type: {path.suffix}")


def parse_month(value: str) -> tuple[int, int]:
    """
    Accept --month 07 (July of current system year) or --month 2026-07.
    """
    text = str(value).strip()
    today = dt.date.today()
    try:
        if re.fullmatch(r"\d{4}-\d{1,2}", text):
            y, m = text.split("-", 1)
            year, month = int(y), int(m)
        elif re.fullmatch(r"\d{1,2}", text):
            year, month = today.year, int(text)
        else:
            raise ValueError
        if not 1 <= month <= 12:
            raise ValueError
        return year, month
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            "Use 07 (Jul of current year) or YYYY-MM, e.g. 2026-08"
        ) from exc


def month_bounds(year: int, month: int) -> tuple[dt.date, dt.date]:
    start = dt.date(year, month, 1)
    if month == 12:
        end = dt.date(year + 1, 1, 1)
    else:
        end = dt.date(year, month + 1, 1)
    return start, end


def saturdays_in_month(year: int, month: int) -> list[dt.date]:
    """All Saturdays in the month, in order."""
    start, end = month_bounds(year, month)
    sats = []
    d = start
    while d < end:
        if d.weekday() == 5:  # Saturday
            sats.append(d)
        d += dt.timedelta(days=1)
    return sats


def off_saturday(year: int, month: int, which: str = "2nd") -> dt.date | None:
    """
    Which single Saturday is the monthly off day.
    which: 1st | 2nd | 3rd | 4th | last  (default 2nd — common one-Saturday-off policy)
    """
    sats = saturdays_in_month(year, month)
    if not sats:
        return None
    key = str(which).strip().lower()
    mapping = {
        "1": 0,
        "1st": 0,
        "first": 0,
        "2": 1,
        "2nd": 1,
        "second": 1,
        "3": 2,
        "3rd": 2,
        "third": 2,
        "4": 3,
        "4th": 3,
        "fourth": 3,
        "last": -1,
    }
    if key not in mapping:
        raise SystemExit(
            f"Invalid --saturday-off '{which}'. Use: 1st, 2nd, 3rd, 4th, or last."
        )
    idx = mapping[key]
    if idx == -1:
        return sats[-1]
    if idx >= len(sats):
        # e.g. ask for 4th Saturday but month only has 3 → use last Saturday
        return sats[-1]
    return sats[idx]


def working_days_count(year: int, month: int, saturday_off: str = "2nd") -> int:
    """
    Working days for the month:
      - All Sundays are off
      - Only ONE Saturday is off (default: 2nd Saturday)
      - Other Saturdays count as working days
      - Mon–Fri always working
    """
    start, end = month_bounds(year, month)
    off_sat = off_saturday(year, month, saturday_off)
    n = 0
    d = start
    while d < end:
        wd = d.weekday()  # Mon=0 ... Sun=6
        if wd == 6:
            # Sunday always off
            pass
        elif wd == 5:
            # Saturday: off only if it is the chosen monthly off Saturday
            if off_sat is None or d != off_sat:
                n += 1
        else:
            n += 1
        d += dt.timedelta(days=1)
    return n


def parse_seconds(value) -> float:
    """Jira Σ Time Spent is usually seconds. Blank -> 0."""
    if value is None:
        return 0.0
    if isinstance(value, float) and pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text.lower() in {"nan", "none", "-"}:
        return 0.0
    # duration text like 2h 30m
    m = re.match(
        r"^(?:(?P<h>\d+(?:\.\d+)?)\s*h)?\s*(?:(?P<m>\d+(?:\.\d+)?)\s*m)?$",
        text,
        re.I,
    )
    if m and (m.group("h") or m.group("m")):
        return float(m.group("h") or 0) * 3600 + float(m.group("m") or 0) * 60
    try:
        return float(text)
    except ValueError:
        return 0.0


def seconds_to_timedelta(seconds: float) -> dt.timedelta:
    return dt.timedelta(seconds=round(float(seconds or 0)))


def in_month(series: pd.Series, start: dt.date, end: dt.date) -> pd.Series:
    d = pd.to_datetime(series, errors="coerce")
    return d.dt.date.ge(start) & d.dt.date.lt(end)


def sanitize_sheet_name(name: str, used: set) -> str:
    clean = re.sub(r"[\[\]\:\*\?/\\]", "", str(name)).strip() or "Unassigned"
    clean = clean[:31]
    base, i = clean, 2
    while clean.lower() in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def autosize(ws):
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(10, min(length + 2, 45))


def write_excel(
    rows: list[dict],
    output: Path,
    month_label: str,
    working_days: int,
    saturday_off_label: str,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Tasks"

    wd_header = f"Working Days (1 Sat off + Sundays)"
    headers = [
        "Issue Key",
        "Summary",
        "Issue Type",
        "Employee",
        "Created Date",
        "Updated Date",
        f"Time Spent ({month_label})",
        wd_header,
        "Avg / Working Day",
        "Status",
        "Priority",
    ]

    ws.append([f"Employee Task Average — {month_label}"])
    ws.append(
        [
            f"Working days = {working_days} "
            f"(all Sundays off; only {saturday_off_label} Saturday off; other Saturdays count as work). "
            "Include: created in month OR created last month and updated in month. "
            "Time comes from Sigma Time Spent on the Jira sheet."
        ]
    )
    ws.append([])
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r["issue_key"],
                r["summary"],
                r["issue_type"],
                r["employee"],
                r["created"],
                r["updated"],
                seconds_to_timedelta(r["seconds"]),
                working_days,
                seconds_to_timedelta(r["avg_working_day"]),
                r["status"],
                r["priority"],
            ]
        )

    for row_idx in range(5, 5 + len(rows)):
        ws.cell(row=row_idx, column=5).number_format = "dd-mmm-yyyy"
        ws.cell(row=row_idx, column=6).number_format = "dd-mmm-yyyy"
        ws.cell(row=row_idx, column=7).number_format = DURATION_FMT
        ws.cell(row=row_idx, column=9).number_format = DURATION_FMT
    autosize(ws)

    used = {"all tasks"}
    by_emp: dict[str, list[dict]] = {}
    for r in rows:
        by_emp.setdefault(r["employee"], []).append(r)

    for employee in sorted(by_emp.keys(), key=str.lower):
        emp_ws = wb.create_sheet(sanitize_sheet_name(employee, used))
        emp_rows = by_emp[employee]
        emp_ws.append([f"{employee} — {month_label}"])
        emp_ws.append([])
        emp_ws.append(headers)
        for i, r in enumerate(emp_rows):
            emp_ws.append(
                [
                    r["issue_key"],
                    r["summary"],
                    r["issue_type"],
                    r["employee"],
                    r["created"],
                    r["updated"],
                    seconds_to_timedelta(r["seconds"]),
                    working_days,
                    seconds_to_timedelta(r["avg_working_day"]),
                    r["status"],
                    r["priority"],
                ]
            )
            excel_row = 4 + i
            emp_ws.cell(row=excel_row, column=5).number_format = "dd-mmm-yyyy"
            emp_ws.cell(row=excel_row, column=6).number_format = "dd-mmm-yyyy"
            emp_ws.cell(row=excel_row, column=7).number_format = DURATION_FMT
            emp_ws.cell(row=excel_row, column=9).number_format = DURATION_FMT

        total_sec = sum(r["seconds"] for r in emp_rows)
        trow = 4 + len(emp_rows) + 1
        emp_ws.cell(row=trow, column=1, value="Total Tasks")
        emp_ws.cell(row=trow, column=2, value=len(emp_rows))
        emp_ws.cell(row=trow + 1, column=1, value=f"Total Time ({month_label})")
        emp_ws.cell(row=trow + 1, column=2, value=seconds_to_timedelta(total_sec))
        emp_ws.cell(row=trow + 1, column=2).number_format = DURATION_FMT
        emp_ws.cell(row=trow + 2, column=1, value=wd_header)
        emp_ws.cell(row=trow + 2, column=2, value=working_days)
        emp_ws.cell(row=trow + 3, column=1, value="Avg Time / Working Day (all tasks)")
        avg = (total_sec / working_days) if working_days else 0.0
        emp_ws.cell(row=trow + 3, column=2, value=seconds_to_timedelta(avg))
        emp_ws.cell(row=trow + 3, column=2).number_format = DURATION_FMT
        emp_ws.cell(row=trow + 4, column=1, value="Avg Time / Task")
        avg_task = (total_sec / len(emp_rows)) if emp_rows else 0.0
        emp_ws.cell(row=trow + 4, column=2, value=seconds_to_timedelta(avg_task))
        emp_ws.cell(row=trow + 4, column=2).number_format = DURATION_FMT
        autosize(emp_ws)

    wb.save(output)


def main():
    parser = argparse.ArgumentParser(
        description="Per-task averages from Jira sheet only (no leave / worklogs / API)."
    )
    parser.add_argument("--input", required=True, type=Path, help="Jira task export .csv/.xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("Employee_Task_Avg_Report.xlsx"),
        help="Output Excel path",
    )
    parser.add_argument(
        "--month",
        type=parse_month,
        default=None,
        help="Target month: 07 or 2026-07 (default: current calendar month)",
    )
    parser.add_argument(
        "--all-tasks",
        action="store_true",
        help="Do not filter by Created/Updated month; use every row in the sheet",
    )
    parser.add_argument(
        "--saturday-off",
        default="2nd",
        help="Which one Saturday is off each month: 1st, 2nd (default), 3rd, 4th, or last",
    )
    args = parser.parse_args()

    if args.month is None:
        today = dt.date.today()
        year, month = today.year, today.month
    else:
        year, month = args.month
    month_label = f"{year:04d}-{month:02d}"
    start, end = month_bounds(year, month)
    working_days = working_days_count(year, month, args.saturday_off)
    off_sat = off_saturday(year, month, args.saturday_off)
    sat_label = f"{args.saturday_off} ({off_sat.isoformat()})" if off_sat else args.saturday_off

    df = read_table(args.input)
    df = normalize_columns(df)

    required = {"Issue Key", "Summary", "Assignee", "Time Spent"}
    missing = required - set(df.columns)
    if missing:
        raise SystemExit(f"Input missing columns: {sorted(missing)}. Found: {list(df.columns)}")

    for col in ["Created Date", "Updated Date", "Status", "Priority", "Issue Type"]:
        if col not in df.columns:
            df[col] = None

    df["Assignee"] = df["Assignee"].fillna("Unassigned").astype(str).str.strip()
    df["Issue Key"] = df["Issue Key"].astype(str).str.strip()
    df["Created Date"] = pd.to_datetime(df["Created Date"], errors="coerce")
    df["Updated Date"] = pd.to_datetime(df["Updated Date"], errors="coerce")
    df["Seconds"] = df["Time Spent"].apply(parse_seconds)

    if not args.all_tasks:
        # Previous month window (e.g. June when --month 07)
        prev_last = start - dt.timedelta(days=1)
        prev_start = prev_last.replace(day=1)

        created = pd.to_datetime(df["Created Date"], errors="coerce")
        updated = pd.to_datetime(df["Updated Date"], errors="coerce")
        c_dates = created.dt.date
        u_dates = updated.dt.date

        created_in_selected = c_dates.notna() & (c_dates >= start) & (c_dates < end)
        created_in_previous = c_dates.notna() & (c_dates >= prev_start) & (c_dates < start)
        updated_in_selected = u_dates.notna() & (u_dates >= start) & (u_dates < end)
        # Include: created in selected month OR (created last month AND updated in selected month)
        df = df[created_in_selected | (created_in_previous & updated_in_selected)].copy()

    df = df[df["Issue Key"].ne("") & df["Issue Key"].str.lower().ne("nan")]

    if df.empty:
        raise SystemExit(
            f"No tasks found for {month_label}. "
            "Try --all-tasks if your sheet is already month-filtered."
        )

    rows = []
    for _, r in df.iterrows():
        sec = float(r["Seconds"] or 0)
        avg = (sec / working_days) if working_days else 0.0
        rows.append(
            {
                "issue_key": r["Issue Key"],
                "summary": r["Summary"],
                "issue_type": r.get("Issue Type") or "",
                "employee": r["Assignee"],
                "created": r["Created Date"],
                "updated": r["Updated Date"],
                "seconds": sec,
                "avg_working_day": avg,
                "status": r.get("Status") or "",
                "priority": r.get("Priority") or "",
            }
        )

    rows.sort(key=lambda x: (str(x["employee"]).lower(), str(x["issue_key"])))
    write_excel(rows, args.output, month_label, working_days, sat_label)

    print(f"Month: {month_label}")
    print(f"Working days: {working_days} (Sundays off; only {sat_label} Saturday off)")
    print(f"Tasks: {len(rows)}")
    print(f"Report written to: {args.output}")
    print(
        "Note: Time Spent is taken from Time Spent on the sheet "
        "(usually lifetime in Jira, not true month-only split)."
    )


if __name__ == "__main__":
    main()
