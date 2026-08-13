#!/usr/bin/env python3
"""
avg_workingHrs_calcultor.py

Estimate each task's hours for a selected month using:

  Historical Working Days  = valid working days between Created and Updated
                             (exclude leave / weekoff / Sunday)
  Historical Work Hours    = THAT TASK's time spent
                             (issue worklogs for the Issue Key, else Jira Σ Time Spent)
  Average Daily Hours      = Historical Work Hours / Historical Working Days
  Current Month Working Days = valid days in overlap with selected month
  Current Month Task Hours = Average Daily Hours × Current Month Working Days

Tempo user/day worklogs are used only for attendance (who worked which day),
NOT as the task's hours. Task hours always come from the specific issue.

Usage:
  python avg_workingHrs_calcultor.py ^
      --jira ".\\Jira sheet 100.xlsx" ^
      --worklogs ".\\Worklogs.xlsx" ^
      --leave ".\\Leave.xlsx" ^
      --output ".\\Jira_avg_task_emp.xlsx" ^
      --month 07
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import warnings
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Monday=0 ... Sunday=6. Default weekoff: Sunday only.
# Saturdays are working unless leave/attendance marks them off.
DEFAULT_WEEKOFF_DAYS: set[int] = {6}

# Attendance codes treated as NON-WORKING
LEAVE_CODES = {
    "ll", "cl", "cl1", "cl2", "al", "l", "leave", "pl", "sl", "el",
    "comoff", "coff", "co", "fh", "holiday", "ho", "hd", "ml", "rh",
    "sh", "sun", "wo", "off", "ab", "absent", "weekoff", "woff", "na",
}

# Attendance codes treated as PRESENT / working
PRESENT_CODES = {
    "av", "os", "wfh", "p", "pr", "present", "w", "od", "tr", "travel",
    "half", "hdw",  # half-day work still counts as a working day presence
}

_DURATION_RE = re.compile(
    r"^(?:(?P<h>\d+(?:\.\d+)?)\s*h)?\s*(?:(?P<m>\d+(?:\.\d+)?)\s*m)?$",
    re.IGNORECASE,
)
_DATE_HEADER_RE = re.compile(
    r"^(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun),\s*(.+)$",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Name / date helpers
# ---------------------------------------------------------------------------

def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip().lower())


def names_match(a: str, b: str) -> bool:
    a, b = normalize_name(a), normalize_name(b)
    if not a or not b:
        return False
    if a == b:
        return True
    if len(a) >= 5 and len(b) >= 5 and (a in b or b in a):
        return True
    ta, tb = a.split(), b.split()
    if len(ta) >= 1 and len(tb) >= 1 and ta[0] == tb[0]:
        if len(ta) == 1 or len(tb) == 1:
            return len(ta[0]) >= 5
        if ta[-1] == tb[-1] or ta[-1] in tb or tb[-1] in ta:
            return True
    return False


def daterange_inclusive(start: dt.date, end: dt.date):
    d = start
    while d <= end:
        yield d
        d += dt.timedelta(days=1)


def parse_month_arg(value: str) -> tuple[int | None, int]:
    """
    --month 07      -> (None, 7)   year resolved later from data / today
    --month 2026-07 -> (2026, 7)
    """
    text = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{1,2}", text):
        y, m = text.split("-", 1)
        year, month = int(y), int(m)
    elif re.fullmatch(r"\d{1,2}", text):
        year, month = None, int(text)
    else:
        raise argparse.ArgumentTypeError("Use 07 or 2026-07")
    if not 1 <= month <= 12:
        raise argparse.ArgumentTypeError("Month must be 01..12")
    return year, month


def month_bounds(year: int, month: int) -> tuple[dt.date, dt.date, dt.date, str]:
    """Return month_start, next_month_start, month_end_inclusive, label."""
    start = dt.date(year, month, 1)
    if month == 12:
        next_start = dt.date(year + 1, 1, 1)
        end_incl = dt.date(year, 12, 31)
    else:
        next_start = dt.date(year, month + 1, 1)
        end_incl = next_start - dt.timedelta(days=1)
    label = start.strftime("%B %Y")
    return start, next_start, end_incl, label


def resolve_reporting_year(
    month_arg: tuple[int | None, int] | None,
    worklog_dates: list[dt.date],
    jira_dates: list[dt.date],
) -> tuple[int, int]:
    """Pick year from explicit arg, else data, else system year."""
    today = dt.date.today()
    if month_arg is None:
        return today.year, today.month

    year_hint, month = month_arg
    if year_hint is not None:
        return year_hint, month

    candidates = [d.year for d in worklog_dates if d.month == month]
    if not candidates:
        candidates = [d.year for d in jira_dates if d.month == month]
    if not candidates:
        candidates = [d.year for d in (worklog_dates + jira_dates)]
    year = max(candidates) if candidates else today.year
    return year, month


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_worklog_hours(value) -> float:
    """
    Parse Tempo-style durations into decimal hours.
      8h -> 8.0
      7h 30m -> 7.5
      45m -> 0.75
      blank -> 0.0
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600.0
    if isinstance(value, dt.time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, (int, float)):
        # Plain numbers in Tempo exports are usually already hours
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text or text.lower() in {"nan", "none", "-", "—"}:
        return 0.0

    m = _DURATION_RE.match(text)
    if m and (m.group("h") or m.group("m")):
        return float(m.group("h") or 0) + float(m.group("m") or 0) / 60.0

    # HH:MM
    hm = re.fullmatch(r"(\d+):(\d{1,2})", text)
    if hm:
        return int(hm.group(1)) + int(hm.group(2)) / 60.0

    try:
        return float(text)
    except ValueError:
        warnings.warn(f"Unparseable worklog duration '{value}' — treated as 0")
        return 0.0


def parse_header_date(col) -> dt.date | None:
    """Parse 'Wed, 01 Jul 2026' style column headers."""
    text = str(col).strip()
    m = _DATE_HEADER_RE.match(text)
    if m:
        text = m.group(1).strip()
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date()


def read_table(path: Path, header=0) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, sheet_name=0, header=header, engine="openpyxl")
    if suffix == ".csv":
        return pd.read_csv(path, header=header)
    raise SystemExit(f"Unsupported file type: {path}")


def parse_jira_datetime(series: pd.Series) -> pd.Series:
    """
    Parse Jira Created/Updated values.
    Supports both day-first (25-06-2026, 29-07-2026) and US (7/29/2026).
    """
    if series.empty:
        return pd.to_datetime(series, errors="coerce")

    # Already datetime-like from Excel
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")

    text = series.astype(str)
    # Prefer day-first when values look like dd-mm-yyyy
    sample = " ".join(text.head(20).tolist())
    dayfirst = bool(re.search(r"\b\d{1,2}[-/]\d{1,2}[-/]\d{4}\b", sample)) and (
        bool(re.search(r"\b([0-3]?\d)[-/]([0-1]?\d)[-/]\d{4}\b", sample))
    )
    # If any day > 12 appears in the first position, force dayfirst
    force_dayfirst = False
    for v in text.head(50):
        m = re.match(r"^\s*(\d{1,2})[-/](\d{1,2})[-/](\d{4})", str(v))
        if m and int(m.group(1)) > 12:
            force_dayfirst = True
            break

    parsed = pd.to_datetime(series, errors="coerce", dayfirst=force_dayfirst or dayfirst)
    # Fallback pass with opposite dayfirst for remaining NaT
    if parsed.isna().any():
        alt = pd.to_datetime(series, errors="coerce", dayfirst=not (force_dayfirst or dayfirst))
        parsed = parsed.fillna(alt)
    return parsed


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[\[\]\:\*\?/\\]", "", str(name)).strip() or "Unassigned"
    clean = clean[:31]
    base, i = clean, 2
    while clean.lower() in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def normalize_jira_columns(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "Issue Key": ["issue key", "key", "issuekey"],
        "Issue Type": ["issue type", "issuetype", "type"],
        "Summary": ["summary"],
        "Assignee": ["assignee"],
        "Status": ["status"],
        "Created": ["created", "created date"],
        "Updated": ["updated", "updated date"],
        "Project Name": ["project name", "project", "projectname"],
        "Priority": ["priority"],
        "Time Spent": [
            "time spent", "σ time spent", "sigma time spent",
            "sum time spent", "sum of time spent",
        ],
    }
    lookup = {v: k for k, vs in aliases.items() for v in vs}
    rename, used = {}, set()
    for col in df.columns:
        key = re.sub(r"\s+", " ", str(col).strip().lower())
        target = lookup.get(key)
        if not target:
            for v, c in lookup.items():
                if v != "date" and (v in key or key in v):
                    target = c
                    break
        if target and target not in used:
            rename[col] = target
            used.add(target)
    return df.rename(columns=rename)


def jira_time_spent_to_hours(value) -> float:
    """
    Jira Σ Time Spent is usually seconds (e.g. 7200 -> 2.0 hours).
    Also accepts duration text like '2h 30m'.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600.0
    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() in {"nan", "none", "-"}:
            return 0.0
        if re.search(r"[hm]", text, re.I):
            return parse_worklog_hours(text)
        try:
            num = float(text.replace(",", ""))
        except ValueError:
            return 0.0
        # Large numbers are seconds; small plain numbers treated as hours
        return num / 3600.0 if num > 48 else num
    if isinstance(value, (int, float)):
        num = float(value)
        return num / 3600.0 if num > 48 else num
    return parse_worklog_hours(value)


def load_jira(path: Path) -> pd.DataFrame:
    df = normalize_jira_columns(read_table(path))
    need = {"Issue Key", "Summary", "Assignee", "Created", "Updated"}
    missing = need - set(df.columns)
    if missing:
        raise SystemExit(f"Jira sheet missing {sorted(missing)}. Found: {list(df.columns)}")

    for c in ["Issue Type", "Status", "Project Name", "Priority", "Time Spent"]:
        if c not in df.columns:
            df[c] = None

    df["Issue Key"] = df["Issue Key"].astype(str).str.strip()
    df["Assignee"] = df["Assignee"].fillna("").astype(str).str.strip()
    df["Created"] = parse_jira_datetime(df["Created"])
    df["Updated"] = parse_jira_datetime(df["Updated"])
    # Task-specific total hours from Jira sheet (fallback when no issue worklogs)
    df["Task Hours"] = df["Time Spent"].apply(jira_time_spent_to_hours)
    df = df[df["Issue Key"].ne("") & df["Issue Key"].str.lower().ne("nan")]
    return df.drop_duplicates(subset=["Issue Key"], keep="first")


def is_tempo_timesheet(df: pd.DataFrame) -> bool:
    cols = [str(c).strip() for c in df.columns]
    has_user = any(c.lower() == "user" for c in cols)
    date_like = sum(1 for c in cols if _DATE_HEADER_RE.match(c))
    return has_user and date_like >= 1


class WorklogIndex:
    """
    - hours_by_emp: employee daily hours (Tempo / attendance)
    - hours_by_issue: task-specific hours from issue-level worklogs (Issue Key)
    """

    def __init__(self):
        self.hours_by_emp: dict[str, dict[dt.date, float]] = defaultdict(dict)
        self.hours_by_issue: dict[str, float] = {}
        self.display_names: dict[str, str] = {}
        self.coverage: tuple[dt.date, dt.date] | None = None
        self.source_note = ""
        self.has_issue_hours = False
        self.unmatched_warned: set[str] = set()

    def add_hours(self, employee: str, day: dt.date, hours: float):
        key = normalize_name(employee)
        if not key:
            return
        self.display_names.setdefault(key, employee.strip())
        self.hours_by_emp[key][day] = self.hours_by_emp[key].get(day, 0.0) + float(hours)

    def add_issue_hours(self, issue_key: str, hours: float):
        key = str(issue_key or "").strip()
        if not key or key.lower() == "nan":
            return
        self.hours_by_issue[key] = self.hours_by_issue.get(key, 0.0) + float(hours)
        self.has_issue_hours = True

    def task_hours(self, issue_key: str) -> float | None:
        """Return task-specific worklog hours if issue worklogs exist for this key."""
        key = str(issue_key or "").strip()
        if not self.has_issue_hours:
            return None
        if key in self.hours_by_issue:
            return float(self.hours_by_issue[key])
        return 0.0

    def finalize(self):
        all_dates: list[dt.date] = []
        for days in self.hours_by_emp.values():
            all_dates.extend(days.keys())
        if all_dates:
            self.coverage = (min(all_dates), max(all_dates))

    def find_key(self, employee: str) -> str | None:
        target = normalize_name(employee)
        if not target:
            return None
        if target in self.hours_by_emp:
            return target
        for key in self.hours_by_emp:
            if names_match(employee, key) or names_match(employee, self.display_names.get(key, "")):
                return key
        return None

    def hours_on(self, employee: str, day: dt.date) -> float | None:
        key = self.find_key(employee)
        if key is None:
            return None
        return float(self.hours_by_emp[key].get(day, 0.0))

    def all_dates(self) -> list[dt.date]:
        out: list[dt.date] = []
        for days in self.hours_by_emp.values():
            out.extend(days.keys())
        return out


def load_worklogs(path: Path) -> WorklogIndex:
    raw = read_table(path)
    idx = WorklogIndex()

    if is_tempo_timesheet(raw):
        user_col = next(c for c in raw.columns if str(c).strip().lower() == "user")
        date_cols: list[tuple[object, dt.date]] = []
        for c in raw.columns:
            day = parse_header_date(c)
            if day is not None and _DATE_HEADER_RE.match(str(c).strip()):
                date_cols.append((c, day))

        if not date_cols:
            raise SystemExit(f"No date columns found in Tempo worklogs: {path}")

        project_cols = [c for c in raw.columns if str(c).strip().lower() == "project"]
        project_col = project_cols[0] if project_cols else None

        # Prefer Project == "Total" rows (already daily totals). Otherwise sum project rows.
        total_rows = []
        project_rows = []
        for _, row in raw.iterrows():
            user = str(row[user_col]).strip() if pd.notna(row[user_col]) else ""
            if not user or user.lower() in {"nan", "none", "summary"}:
                continue
            is_total = (
                project_col is not None
                and pd.notna(row[project_col])
                and str(row[project_col]).strip().lower() == "total"
            )
            if is_total:
                total_rows.append(row)
            else:
                project_rows.append(row)

        rows_to_use = total_rows if total_rows else project_rows
        for row in rows_to_use:
            user = str(row[user_col]).strip()
            for col, day in date_cols:
                hrs = parse_worklog_hours(row[col])
                if hrs > 0:
                    idx.add_hours(user, day, hrs)

        idx.source_note = (
            "Tempo user/day timesheet (attendance only; "
            "task hours from Jira Sigma Time Spent)"
        )
        idx.finalize()
        return idx

    # Issue-level worklogs: hours by Issue Key (task-specific) + Author/day (attendance)
    aliases = {
        "Issue Key": ["issue key", "key", "issuekey", "issue", "issue id"],
        "Author": ["author", "user", "worker", "assignee", "display name", "full name"],
        "Started": ["started", "work date", "worklog date", "date", "start date", "tempo date"],
        "Time Spent": ["time spent", "hours", "duration", "spent", "time spent (seconds)", "timespentseconds"],
    }
    lookup = {v: k for k, vs in aliases.items() for v in vs}
    rename, used = {}, set()
    for col in raw.columns:
        key = re.sub(r"\s+", " ", str(col).strip().lower())
        target = lookup.get(key)
        if not target:
            for v, k in lookup.items():
                if v in key:
                    target = k
                    break
        if target and target not in used:
            rename[col] = target
            used.add(target)
    wl = raw.rename(columns=rename)

    if "Started" not in wl.columns:
        raise SystemExit(
            f"Worklogs must be Tempo (User + daily columns) or have Date/Started. "
            f"Found: {list(raw.columns)}"
        )
    if "Time Spent" not in wl.columns:
        raise SystemExit("Worklogs missing Time Spent column.")

    if "Author" not in wl.columns:
        wl["Author"] = ""
    else:
        wl["Author"] = wl["Author"].fillna("").astype(str).str.strip()
    wl["Started"] = pd.to_datetime(wl["Started"], errors="coerce")

    def to_hours(v):
        if isinstance(v, (int, float)) and not pd.isna(v) and float(v) > 48:
            return float(v) / 3600.0  # likely seconds
        return parse_worklog_hours(v)

    wl["Hours"] = wl["Time Spent"].apply(to_hours)
    wl = wl[wl["Started"].notna() & (wl["Hours"] > 0)]

    if "Issue Key" in wl.columns:
        wl["Issue Key"] = wl["Issue Key"].astype(str).str.strip()
        for _, row in wl.iterrows():
            idx.add_issue_hours(row["Issue Key"], float(row["Hours"]))
            if row["Author"]:
                idx.add_hours(row["Author"], row["Started"].date(), float(row["Hours"]))
        idx.source_note = "Issue-level worklogs (task hours by Issue Key)"
    else:
        for _, row in wl.iterrows():
            if row["Author"]:
                idx.add_hours(row["Author"], row["Started"].date(), float(row["Hours"]))
        idx.source_note = (
            "Worklogs without Issue Key (attendance only; "
            "task hours from Jira Sigma Time Spent)"
        )

    idx.finalize()
    return idx


class LeaveIndex:
    """Employee non-working dates from attendance calendar or flat leave table."""

    def __init__(self):
        # normalized_name -> {date: code}
        self.status_by_emp: dict[str, dict[dt.date, str]] = defaultdict(dict)
        self.non_working: dict[str, set[dt.date]] = defaultdict(set)
        self.holidays: set[dt.date] = set()
        self.display_names: dict[str, str] = {}
        self.sheet_month: tuple[int, int] | None = None  # (year, month) if known
        self.source_note = ""

    def find_key(self, employee: str) -> str | None:
        target = normalize_name(employee)
        if not target:
            return None
        if target in self.status_by_emp or target in self.non_working:
            return target
        for key in set(self.status_by_emp) | set(self.non_working):
            if names_match(employee, key) or names_match(employee, self.display_names.get(key, "")):
                return key
        return None

    def code_on(self, employee: str, day: dt.date) -> str | None:
        key = self.find_key(employee)
        if key is None:
            return None
        return self.status_by_emp.get(key, {}).get(day)

    def is_non_working_status(self, employee: str, day: dt.date) -> bool | None:
        """
        True  -> explicitly non-working
        False -> explicitly present
        None  -> no leave info for this employee/date
        """
        key = self.find_key(employee)
        if key is None:
            return None
        if day in self.non_working.get(key, set()):
            return True
        code = self.status_by_emp.get(key, {}).get(day)
        if code is None:
            return None
        code_l = code.lower()
        if code_l in PRESENT_CODES:
            return False
        if code_l in LEAVE_CODES:
            return True
        # Unknown non-blank code: treat as non-working to be safe
        if code_l:
            return True
        return None


def load_leave(path: Path, year: int, month: int) -> LeaveIndex:
    idx = LeaveIndex()
    idx.sheet_month = (year, month)

    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(path, header=None)
    else:
        raw = pd.read_excel(path, sheet_name=0, header=None, engine="openpyxl")

    header_row = None
    daynum_row = None
    for i in range(min(15, len(raw))):
        vals = [str(v).strip().lower() if pd.notna(v) else "" for v in raw.iloc[i].tolist()]
        if any(v in {"employee name", "employee"} for v in vals):
            header_row = i
            if i + 1 < len(raw):
                daynum_row = i + 1
            break

    if header_row is not None and daynum_row is not None:
        day_numbers: list[int | None] = []
        for v in raw.iloc[daynum_row].tolist()[1:]:
            try:
                day_numbers.append(None if pd.isna(v) else int(float(v)))
            except (TypeError, ValueError):
                day_numbers.append(None)

        for i in range(daynum_row + 1, len(raw)):
            row = raw.iloc[i].tolist()
            name = row[0]
            if pd.isna(name) or str(name).strip() == "":
                continue
            name = str(name).strip()
            name_l = normalize_name(name)
            is_holiday_row = "holiday" in name_l

            for col_offset, day_num in enumerate(day_numbers):
                if day_num is None or day_num < 1 or day_num > 31:
                    continue
                try:
                    d = dt.date(year, month, day_num)
                except ValueError:
                    continue
                col_idx = col_offset + 1
                if col_idx >= len(row):
                    continue
                code = row[col_idx]
                if pd.isna(code) or str(code).strip() == "":
                    # Blank attendance cell → non-working for that employee
                    if not is_holiday_row:
                        idx.non_working[name_l].add(d)
                        idx.status_by_emp[name_l][d] = ""
                        idx.display_names[name_l] = name
                    continue

                code_s = str(code).strip()
                code_l = code_s.lower()
                if is_holiday_row or code_l in {"holiday", "ho", "fh"}:
                    idx.holidays.add(d)

                if is_holiday_row:
                    continue

                idx.status_by_emp[name_l][d] = code_s
                idx.display_names[name_l] = name
                if code_l in PRESENT_CODES:
                    continue
                # Leave / weekoff / Sunday marker / unknown → non-working
                idx.non_working[name_l].add(d)

        idx.source_note = (
            f"Attendance calendar for {year:04d}-{month:02d} "
            f"({len(idx.display_names)} employees)"
        )
        return idx

    # Flat leave table fallback
    df = read_table(path)
    cols = {re.sub(r"\s+", " ", str(c).strip().lower()): c for c in df.columns}
    emp_col = next(
        (cols[k] for k in ("employee", "employee name", "name", "assignee", "user") if k in cols),
        None,
    )
    leave_col = next((cols[k] for k in ("leave date", "date", "leave_date") if k in cols), None)
    from_col = next((cols[k] for k in ("from date", "from", "start date", "start") if k in cols), None)
    to_col = next((cols[k] for k in ("to date", "to", "end date", "end") if k in cols), None)
    type_col = next((cols[k] for k in ("leave type", "type", "category", "code") if k in cols), None)

    if emp_col is None:
        raise SystemExit(
            f"Leave sheet not recognized (need Employee Name calendar or Employee+Date). "
            f"Found: {list(df.columns)}"
        )

    for _, row in df.iterrows():
        emp = str(row[emp_col]).strip()
        if not emp or emp.lower() in {"nan", "none"}:
            continue
        key = normalize_name(emp)
        idx.display_names[key] = emp
        ltype = str(row[type_col]).strip().lower() if type_col and pd.notna(row.get(type_col)) else ""
        is_holiday = "holiday" in ltype or "holiday" in key

        dates: list[dt.date] = []
        if leave_col and pd.notna(row.get(leave_col)):
            d = pd.to_datetime(row[leave_col], errors="coerce")
            if pd.notna(d):
                dates.append(d.date())
        if from_col and pd.notna(row.get(from_col)):
            d0 = pd.to_datetime(row[from_col], errors="coerce")
            d1 = pd.to_datetime(row[to_col], errors="coerce") if to_col else d0
            if pd.notna(d0):
                if pd.isna(d1):
                    d1 = d0
                dates.extend(daterange_inclusive(d0.date(), d1.date()))

        for d in dates:
            if is_holiday:
                idx.holidays.add(d)
            else:
                idx.non_working[key].add(d)
                idx.status_by_emp[key][d] = ltype or "leave"

    idx.source_note = "Flat leave/holiday table"
    return idx


# ---------------------------------------------------------------------------
# Working-day logic
# ---------------------------------------------------------------------------

def is_working_day(
    employee: str,
    day: dt.date,
    *,
    leave: LeaveIndex,
    worklogs: WorklogIndex,
    weekoff_days: set[int] = DEFAULT_WEEKOFF_DAYS,
) -> bool:
    """
    True only when the employee is expected to work that day.

    False for:
      - configured weekoffs (default Sunday)
      - company holidays
      - employee leave / absent / weekoff / other non-working status

    Leave present codes (OS/AV/...) count as working.
    Saturdays are working unless leave marks them off.
    Tempo worklogs are optional attendance hints only (not required for task hours).
    """
    if day.weekday() in weekoff_days:
        return False
    if day in leave.holidays:
        return False

    leave_flag = leave.is_non_working_status(employee, day)
    if leave_flag is True:
        return False
    if leave_flag is False:
        return True

    # No leave row for this date: if Tempo covers the day and employee logged 0,
    # treat as non-working; otherwise count as working (Sat allowed).
    emp_key = worklogs.find_key(employee)
    if (
        emp_key is not None
        and worklogs.coverage is not None
        and worklogs.coverage[0] <= day <= worklogs.coverage[1]
    ):
        hours = float(worklogs.hours_by_emp[emp_key].get(day, 0.0))
        return hours > 0

    return True


def count_working_days(
    employee: str,
    start: dt.date,
    end: dt.date,
    *,
    leave: LeaveIndex,
    worklogs: WorklogIndex,
    weekoff_days: set[int] = DEFAULT_WEEKOFF_DAYS,
) -> list[dt.date]:
    if start is None or end is None or start > end:
        return []
    return [
        d
        for d in daterange_inclusive(start, end)
        if is_working_day(
            employee, d, leave=leave, worklogs=worklogs, weekoff_days=weekoff_days
        )
    ]


def task_in_selected_month(
    created: dt.date | None,
    updated: dt.date | None,
    month_start: dt.date,
    next_month_start: dt.date,
) -> bool:
    """
    Include a task for --month 2026-07 (July) when:

      Created is in July  OR  Updated is in July

    Examples for 2026-07:
      Created 2026-06, Updated 2026-07 → INCLUDE
      Created 2026-07, Updated 2026-07 → INCLUDE
      Created 2026-06, Updated 2026-06 → EXCLUDE
      Created 2026-08, Updated 2026-08 → EXCLUDE
    """
    created_in = created is not None and month_start <= created < next_month_start
    updated_in = updated is not None and month_start <= updated < next_month_start
    return created_in or updated_in


def resolve_task_hours(
    issue_key: str,
    jira_task_hours: float,
    worklogs: WorklogIndex,
) -> tuple[float, str]:
    """
    Task-specific hours only:
      1) Sum of issue-level worklogs for this Issue Key (preferred)
      2) Else Jira Σ Time Spent for this task
    Never uses employee Tempo daily totals (those are attendance, not per-task).
    """
    issue_hrs = worklogs.task_hours(issue_key)
    if issue_hrs is not None and worklogs.has_issue_hours:
        if issue_hrs > 0:
            return float(issue_hrs), "Issue worklogs (this task)"
        # Issue worklogs present but none for this key → fall back to Jira
        if jira_task_hours > 0:
            return float(jira_task_hours), "Jira Sigma Time Spent (no issue worklogs for key)"
        return 0.0, "No issue worklogs and no Jira Sigma Time Spent"

    if jira_task_hours > 0:
        return float(jira_task_hours), "Jira Sigma Time Spent (this task)"
    return 0.0, "No task hours on Jira sheet"


def calculate_task_row(
    row: pd.Series,
    *,
    month_start: dt.date,
    next_month_start: dt.date,
    month_end: dt.date,
    leave: LeaveIndex,
    worklogs: WorklogIndex,
    weekoff_days: set[int],
) -> dict | None:
    """
    Return output dict for one eligible task, or None if not eligible.

    Formula:
      task_hours = this issue's time (worklogs or Jira Σ Time Spent)
      avg_daily  = task_hours / historical_working_days
      month_hrs  = avg_daily * current_month_working_days
    """
    created_ts = row["Created"]
    updated_ts = row["Updated"]
    assignee = str(row.get("Assignee") or "").strip()
    issue_key = str(row.get("Issue Key") or "").strip()

    if pd.isna(created_ts) or pd.isna(updated_ts):
        warnings.warn(
            f"{issue_key or '?'}: missing Created/Updated — skipped"
        )
        return None
    if not assignee:
        warnings.warn(f"{issue_key or '?'}: empty Assignee — skipped")
        return None

    created = created_ts.date()
    updated = updated_ts.date()
    if updated < created:
        updated = created

    if not task_in_selected_month(created, updated, month_start, next_month_start):
        return None

    notes: list[str] = []

    if worklogs.find_key(assignee) is None and worklogs.hours_by_emp:
        if assignee not in worklogs.unmatched_warned:
            warnings.warn(
                f"Employee not found in Worklogs attendance: '{assignee}' "
                "(working days use Leave/Sunday rules only)"
            )
            worklogs.unmatched_warned.add(assignee)

    if leave.find_key(assignee) is None:
        notes.append("Employee not found in Leave sheet — using Sunday/weekoff rules only")

    hist_days = count_working_days(
        assignee, created, updated,
        leave=leave, worklogs=worklogs, weekoff_days=weekoff_days,
    )

    jira_hours = float(row.get("Task Hours") or 0.0)
    hist_hours, hours_source = resolve_task_hours(issue_key, jira_hours, worklogs)
    notes.append(f"Hours source: {hours_source}")

    cur_start = max(created, month_start)
    cur_end = min(updated, month_end)
    cur_days = count_working_days(
        assignee, cur_start, cur_end,
        leave=leave, worklogs=worklogs, weekoff_days=weekoff_days,
    )

    unable = False
    if hist_hours <= 0:
        avg_daily = None
        month_hours = None
        notes.append("Unable to calculate — no task-specific hours")
        unable = True
    elif len(hist_days) == 0:
        avg_daily = None
        month_hours = None
        notes.append("Unable to calculate — zero historical working days")
        unable = True
    else:
        # workinghrs / working days = daily working hrs
        avg_daily = hist_hours / len(hist_days)
        # daily * current month working days = that month working hrs
        month_hours = avg_daily * len(cur_days)

    return {
        "Issue Type": row.get("Issue Type") or "",
        "Issue Key": issue_key,
        "Summary": row.get("Summary") or "",
        "Assignee": assignee,
        "Status": row.get("Status") or "",
        "Created": created,
        "Updated": updated,
        "Project Name": row.get("Project Name") or "",
        "Priority": row.get("Priority") or "",
        "Historical Working Days": len(hist_days),
        "Historical Work Hours": None if unable and hist_hours <= 0 else round(hist_hours, 4),
        "Average Daily Hours": None if avg_daily is None else round(avg_daily, 4),
        "Current Month Working Days": len(cur_days),
        "Current Month Task Hours": None if month_hours is None else round(month_hours, 2),
        "Current Month Task Time": format_hours_clock(month_hours),
        "Notes": "; ".join(notes),
        "_month_hours_raw": month_hours,
        "_unable": unable,
    }


# ---------------------------------------------------------------------------
# Excel output — one sheet per employee
# ---------------------------------------------------------------------------

OUTPUT_COLUMNS = [
    "Issue Type",
    "Issue Key",
    "Summary",
    "Assignee",
    "Status",
    "Created",
    "Updated",
    "Project Name",
    "Priority",
    "Historical Working Days",
    "Historical Work Hours",
    "Average Daily Hours",
    "Current Month Working Days",
    "Current Month Task Hours",
    "Current Month Task Time",
    "Notes",
]


def format_hours_clock(hours) -> str:
    """
    Convert decimal hours to clock text.
      0.81 → 0h 48m 36s   (0.81 × 3600 = 2916 seconds)
      2.5  → 2h 30m
      None → blank
    """
    if hours is None:
        return ""
    try:
        hval = float(hours)
    except (TypeError, ValueError):
        return ""
    if hval < 0 or (isinstance(hval, float) and pd.isna(hval)):
        return ""

    total_seconds = int(round(hval * 3600))
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    if s:
        return f"{h}h {m}m {s}s"
    return f"{h}h {m}m"


def format_date_display(value) -> str:
    """Readable date text for Excel (avoids serial numbers like 46178)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, dt.date):
        return value.strftime("%d-%b-%Y")
    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none"}:
        return ""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.notna(parsed):
        return parsed.strftime("%d-%b-%Y")
    return text


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 48))
        ws.column_dimensions[letter].width = width


def write_employee_workbook(
    rows: list[dict],
    output: Path,
    month_label: str,
):
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    by_emp: dict[str, list[dict]] = {}
    for r in rows:
        by_emp.setdefault(r["Assignee"], []).append(r)

    for employee in sorted(by_emp.keys(), key=lambda x: str(x).lower()):
        emp_rows = sorted(
            by_emp[employee],
            key=lambda r: (r["Created"] or dt.date.min, r["Issue Key"]),
        )
        ws = wb.create_sheet(sanitize_sheet_name(employee, used_names))
        ws.append([f"{employee} - Current Month Task Time ({month_label})"])
        ws.append([
            "Current Month Task Hours = Average Daily Hours x Current Month Working Days. "
            "Current Month Task Time = same value as clock time (e.g. 0.81 hrs = 0h 48m 36s). "
            "Average Daily = THIS TASK's hours / Historical Working Days "
            "(Created to Updated, excluding leave/weekoff/Sunday). "
            "Task hours = issue worklogs for that Issue Key, else Jira Sigma Time Spent."
        ])
        ws.append([])
        ws.append(OUTPUT_COLUMNS)

        for r in emp_rows:
            out_row = []
            for c in OUTPUT_COLUMNS:
                val = r.get(c)
                if c in ("Created", "Updated"):
                    # Write as text so Excel shows 05-Jun-2026, not 46178
                    out_row.append(format_date_display(val))
                else:
                    out_row.append(val)
            ws.append(out_row)

        first = 5
        last = 4 + len(emp_rows)
        # Current Month Task Time is column 15 (after Current Month Task Hours)
        time_col = OUTPUT_COLUMNS.index("Current Month Task Time") + 1
        right_align = Alignment(horizontal="right", vertical="center")
        # Header + data rows
        for row_idx in range(4, last + 1):
            ws.cell(row=row_idx, column=time_col).alignment = right_align
        for row_idx in range(first, last + 1):
            # Keep Created/Updated as plain text (no date serial formatting)
            for col in (11, 12, 14):
                ws.cell(row=row_idx, column=col).number_format = "0.00"

        # Totals
        trow = last + 2
        ws.cell(row=trow, column=1, value="Total Tasks")
        ws.cell(row=trow, column=2, value=len(emp_rows))

        calculable = [
            r["_month_hours_raw"]
            for r in emp_rows
            if r.get("_month_hours_raw") is not None and not r.get("_unable")
        ]
        total_hours = sum(calculable) if calculable else 0.0
        ws.cell(row=trow + 1, column=1, value=f"Total Current Month Task Hours ({month_label})")
        ws.cell(row=trow + 1, column=2, value=round(total_hours, 2))
        ws.cell(row=trow + 1, column=2).number_format = "0.00"

        unable_n = sum(1 for r in emp_rows if r.get("_unable"))
        if unable_n:
            ws.cell(row=trow + 2, column=1, value="Tasks unable to calculate")
            ws.cell(row=trow + 2, column=2, value=unable_n)

        autosize(ws)
        ws.freeze_panes = "A5"

    wb.save(output)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Per-employee current-month task hours from Jira + Worklogs + Leave. "
            "Task hours = that issue's time (Jira Sigma Time Spent or issue worklogs). "
            "Month hours = (task_hours / working_days) × current_month_working_days."
        )
    )
    parser.add_argument("--jira", required=True, type=Path, help="Jira task Excel/CSV")
    parser.add_argument("--worklogs", required=True, type=Path, help="Worklogs / Tempo timesheet Excel")
    parser.add_argument("--leave", required=True, type=Path, help="Employee leave/attendance Excel")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("Jira_avg_task_emp.xlsx"),
        help="Output workbook (one sheet per employee)",
    )
    parser.add_argument(
        "--month",
        type=parse_month_arg,
        default=None,
        help=(
            "Selected month only: 07, 08, or 2026-07 (default = current month). "
            "Includes tasks Created OR Updated in that month "
            "(e.g. Created 2026-06 + Updated 2026-07 is included for 2026-07; "
            "June-only or August-only tasks are excluded). "
            "Does not change the avg-daily hour calculation."
        ),
    )
    parser.add_argument(
        "--weekoff",
        default="sun",
        help="Weekoff weekdays, e.g. sun or sat,sun (default: sun)",
    )
    args = parser.parse_args()

    # Parse weekoffs
    mapping = {
        "mon": 0, "monday": 0, "tue": 1, "tuesday": 1, "wed": 2, "wednesday": 2,
        "thu": 3, "thursday": 3, "fri": 4, "friday": 4,
        "sat": 5, "saturday": 5, "sun": 6, "sunday": 6,
    }
    weekoff: set[int] = set()
    for part in str(args.weekoff).replace(" ", "").lower().split(","):
        if part not in mapping:
            raise SystemExit(f"Unknown weekoff '{part}'. Use mon..sun")
        weekoff.add(mapping[part])
    if not weekoff:
        weekoff = set(DEFAULT_WEEKOFF_DAYS)

    # Load inputs (worklogs first so year can be inferred)
    jira = load_jira(args.jira)
    worklogs = load_worklogs(args.worklogs)

    jira_dates = []
    for col in ("Created", "Updated"):
        for v in jira[col].dropna():
            jira_dates.append(pd.Timestamp(v).date())

    year, month = resolve_reporting_year(args.month, worklogs.all_dates(), jira_dates)
    month_start, next_month_start, month_end, month_label = month_bounds(year, month)

    leave = load_leave(args.leave, year, month)

    rows: list[dict] = []
    excluded = 0
    for _, row in jira.iterrows():
        result = calculate_task_row(
            row,
            month_start=month_start,
            next_month_start=next_month_start,
            month_end=month_end,
            leave=leave,
            worklogs=worklogs,
            weekoff_days=weekoff,
        )
        if result is None:
            excluded += 1
            continue
        rows.append(result)

    if not rows:
        raise SystemExit(
            f"No tasks Created or Updated in {month_label} "
            f"({month_start} .. {month_end}). "
            f"Read {len(jira)}, excluded {excluded}."
        )

    write_employee_workbook(rows, args.output, month_label)

    calculable = [r for r in rows if not r.get("_unable") and r.get("_month_hours_raw") is not None]
    total_month = sum(r["_month_hours_raw"] for r in calculable)
    unable_n = sum(1 for r in rows if r.get("_unable"))

    print(f"Selected Month: {month_label}")
    print(f"Month window: {month_start} .. {month_end}")
    print(
        f"Task filter: Created in {month_label} OR Updated in {month_label} "
        f"(e.g. Created Jun + Updated Jul -> included; other months excluded)"
    )
    print(f"Weekoffs: {sorted(weekoff)} (0=Mon .. 6=Sun)")
    print(f"Worklog source: {worklogs.source_note}")
    print(f"Leave source: {leave.source_note}")
    print()
    print(f"Total Jira Tasks Read: {len(jira)}")
    print(f"Tasks Included: {len(rows)}")
    print(f"Tasks Excluded: {excluded}")
    print(f"Employees: {len({r['Assignee'] for r in rows})}")
    print(f"Tasks unable to calculate: {unable_n}")
    print(f"Total Current Month Task Hours: {total_month:.2f}")
    print()
    print(f"Output:\n{args.output}")


if __name__ == "__main__":
    main()
